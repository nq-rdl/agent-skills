"""Convert Excel workbook sheets to pipe-delimited CSV files (typer CLI version)."""

from pathlib import Path

import openpyxl
import pyarrow as pa
import pyarrow.csv as pcsv
import typer
from rich.console import Console

console = Console()
app = typer.Typer(add_completion=False)


def sanitize_filename(filename: str) -> str:
    """Remove hyphens and replace spaces with underscores."""
    return filename.replace("-", "").replace(" ", "_")


def excel_sheet_to_arrow(workbook, sheet_name: str) -> pa.Table:
    """Convert one Excel sheet to a PyArrow Table."""
    ws = workbook[sheet_name]
    data = [row for row in ws.iter_rows(values_only=True)]
    if not data:
        return pa.table({})

    headers = [f"Column_{i}" if h is None else str(h) for i, h in enumerate(data[0])]
    rows = data[1:]

    arrays = []
    for col_idx, header in enumerate(headers):
        col_data = [
            row[col_idx] if row and col_idx < len(row) else None for row in rows
        ]
        try:
            arr = pa.array(col_data)
        except Exception:
            arr = pa.array([str(v) if v is not None else None for v in col_data])
        arrays.append(arr)

    return pa.table(arrays, names=headers)


@app.command()
def main(
    excel_path: Path = typer.Argument(..., help="Path to the Excel workbook (.xlsx)"),
    output_dir: Path | None = typer.Option(
        None,
        "--output-dir",
        "-o",
        help="Output directory (default: same as Excel file)",
    ),
    prefix: str | None = typer.Option(
        None,
        "--prefix",
        "-p",
        help="Only process sheets starting with this prefix (e.g. 'E.')",
    ),
    force: bool = typer.Option(
        False, "--force/--no-force", "-f", help="Overwrite existing CSV files"
    ),
    cleanup: bool = typer.Option(
        False,
        "--cleanup/--no-cleanup",
        help="Remove existing CSVs and Excel files in output dir first",
    ),
) -> None:
    """Convert all sheets in an Excel workbook to pipe-delimited CSV files."""
    if not excel_path.exists():
        console.print(f"[red]File not found:[/red] {excel_path}")
        raise typer.Exit(1)
    if excel_path.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
        console.print(f"[red]Not an Excel file:[/red] {excel_path}")
        raise typer.Exit(1)

    out_dir = output_dir or excel_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    if cleanup:
        removed = 0
        for pat in ("*.csv", "*.xlsx", "*.xls", "*.xlsm"):
            for f in out_dir.glob(pat):
                f.unlink()
                removed += 1
        console.print(f"[dim]Removed {removed} file(s) from {out_dir}[/dim]")

    base = sanitize_filename(excel_path.stem)

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        console.print(f"[red]Failed to open workbook:[/red] {e}")
        raise typer.Exit(1)

    sheet_names = wb.sheetnames
    if prefix:
        sheet_names = [s for s in sheet_names if s.startswith(prefix)]

    console.print(
        f"Processing [bold]{len(sheet_names)}[/bold] sheet(s) from [bold]{excel_path.name}[/bold]"
    )

    write_opts = pcsv.WriteOptions(delimiter="|")
    ok = 0
    for sheet_name in sheet_names:
        safe_sheet = sanitize_filename(sheet_name)
        dest = out_dir / f"{base}_{safe_sheet}.csv"

        if dest.exists() and not force:
            console.print(
                f"  [yellow]Skip[/yellow] {dest.name} (exists; use --force to overwrite)"
            )
            continue

        try:
            tbl = excel_sheet_to_arrow(wb, sheet_name)
            pcsv.write_csv(tbl, dest, write_opts)
            console.print(
                f"  [green]✓[/green] {dest.name}  ({tbl.num_rows} rows, {len(tbl.column_names)} cols)"
            )
            ok += 1
        except Exception as e:
            console.print(f"  [red]✗ {sheet_name}:[/red] {e}")

    wb.close()
    console.print(f"\n[bold green]{ok}[/bold green] sheet(s) written to {out_dir}")


if __name__ == "__main__":
    app()
